import os
import subprocess
import time

from ichrome import AsyncChrome
import asyncio
from bs4 import BeautifulSoup
import keyboard
import shutil
from openpyxl import load_workbook
from utils import find_path, is_port_used


def cp_excel() -> str:
    """
    复制excel模板
    :return: 文件路径, 复制失败返回空字符串。
    """
    base_path = os.path.dirname(__file__)
    tmp_path = os.path.join(base_path, "excel_template", "tmp.xlsx")
    today = time.strftime("%Y%m%d", time.localtime())
    result_path = os.path.join(base_path, "outputs", f"{today}.xlsx")
    shutil.copyfile(tmp_path, result_path)
    if os.path.exists(result_path):
        print("复制成功")
        return result_path
    else:
        print("复制失败")
        return ""


# excel追加写入（xlsx：批注不会丢失）
def write_excel_xlsx_append(file_path, title="", url="", author=""):
    """
    追加写入excel
    :param file_path: 文件保存路径
    :param title: 文章标题
    :param url: 文章链接
    :param author: 文章作者
    """
    wb = load_workbook(file_path)
    ws = wb.active
    used_rows = 1

    while True:
        if ws[f"A{used_rows}"].value is None:
            break
        used_rows += 1

    ws[f"A{used_rows}"] = used_rows - 1     # 序号
    ws[f"B{used_rows}"] = title             # 标题
    ws[f"C{used_rows}"] = url               # 链接
    ws[f"D{used_rows}"] = author            # 作者
    wb.save(file_path)


def run_debug_mode() -> bool:
    """
    调试模式启动chrome浏览器，并自动进入InfoQ写作社区
    :return:
    """
    cmd_str = ""
    browser_path = find_path("chrome.exe")
    if browser_path != "":
        cmd_str = f'"{browser_path}" --remote-debugging-port=9222 --remote-allow-origins=* https://xie.infoq.cn/latest'
    if cmd_str == "":
        print("未找到浏览器")
        return False
    subprocess.Popen(cmd_str, shell=True)
    while not is_port_used():
        time.sleep(1)
    return True


async def main(fp):
    """
    获取当前页面的标题、链接、作者，并写入excel, 最后弹出提示框
    :param fp:
    :return:
    """
    async with AsyncChrome() as chrome:
        async with chrome.connect_tab() as tab:

            print(await tab.title)
            print(await tab.url)
            soup = BeautifulSoup(await tab.html, 'html.parser')
            print(soup.find("a", class_="com-author-name").text)
            write_excel_xlsx_append(fp,
                                    title=await tab.title,
                                    url=await tab.url,
                                    author=soup.find("a", class_="com-author-name").text
                                    )

            await tab.alert(
                '数据已保存在excel中，按Esc退出程序'
            )


def run(fp):
    asyncio.run(main(fp))


if __name__ == "__main__":
    if run_debug_mode():
        print("调试模式启动成功")
        fp = cp_excel()
        keyboard.add_hotkey('q', run, args=(fp,))
        rec = keyboard.record(until='esc')
        # print(rec)
